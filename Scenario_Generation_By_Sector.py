import pandas as pd
import openpyxl

sectors = ["电力", "工业", "交通", "民用", "农业"]
wb = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\early_peak-net_zero-clean_air\2020_2035_change_percentage.xlsx")

for i, sector in enumerate(sectors):
    for j, sector_2 in enumerate(sectors):
        for z, sector_3 in enumerate(sectors):
            if i >= j or j >= z or i >= z:
                continue

            wb_edit = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\Emission Excel Files\管控方案模板.xlsx")
            ws_edit = wb_edit["管控方案"]
            ws_edit.delete_rows(3, 10000)
            for sheet_name in ["SO2", "NOX", "VOC", "NH3", "PM2.5"]:
                ws = wb[sheet_name]
                for r in range (2, 7):
                    for c in range(2, 33):
                        if (r == i + 2 or r == j + 2 or r == z + 2):
                            province = ws.cell(row = 1, column = c).value
                            value = ws.cell(row = r, column = c).value

                            if value == "#DIV/0!":
                                value = 0
                            elif float(value) < 0:
                                value = 0
                
                            if sheet_name == "PM2.5":
                                sheet_name = "PM"

                            if r == i + 2:
                                ws_edit.append([province, sector, sheet_name, value])
                            elif r == j + 2:
                                ws_edit.append([province, sector_2, sheet_name, value])
                            else:
                                ws_edit.append([province, sector_3, sheet_name, value])

                for weather_year in range(2017, 2025):
                    wb_edit.save(rf"C:\Users\sunyi\Environment_Project\Scenario Excel Files Sector\Early-Peak-{sector}-{sector_2}-{sector_3}.xlsx")
    
print("Finished!")