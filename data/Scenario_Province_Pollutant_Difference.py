'''
For each pollutant, renders a 5x5 choropleth grid (sector x
scenario) of Chinese provinces showing absolute 2020->2035 emission
change, with national sector totals labeled, and saves each figure
as a PNG.
'''

import cartopy.crs as ccrs
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.colors as mcolors
from common import PROVINCE_MAP_EMISSIONS, PROVINCE_ARR, EXCLUDED_REGIONS, get_province_gdf

wb_baseline = openpyxl.load_workbook("baseline/2020_2035_Raw.xlsx")
wb_clean_air = openpyxl.load_workbook("clean_air/2020_2035_Raw.xlsx")
wb_otpca = openpyxl.load_workbook("on-time_peak-clean_air/2020_2035_Raw.xlsx")
wb_otpnzca = openpyxl.load_workbook("on-time_peak-net_zero-clean_air/2020_2035_Raw.xlsx")
wb_epnzca = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_2035_Raw.xlsx")

workbook_arr = [wb_baseline, wb_clean_air, wb_otpca, wb_otpnzca, wb_epnzca]

pollutant_norms = {
    "SO2":   mcolors.SymLogNorm(linthresh=10000, vmin=-520000, vmax=200000),
    "NOX":   mcolors.SymLogNorm(linthresh=10000, vmin=-400000, vmax=180000),
    "VOC":   mcolors.SymLogNorm(linthresh=10000, vmin=-660000, vmax=260000),
    "NH3":   mcolors.SymLogNorm(linthresh=1000,  vmin=-180000, vmax=30000),
    "PM2.5": mcolors.SymLogNorm(linthresh=5000,  vmin=-310000, vmax=110000),
    "PM10":  mcolors.SymLogNorm(linthresh=5000,  vmin=-380000, vmax=140000),
    "BC":    mcolors.SymLogNorm(linthresh=1000,  vmin=-80000,  vmax=30000),
    "OC":    mcolors.SymLogNorm(linthresh=1000,  vmin=-160000, vmax=60000),
}

pollutant_ticks = {
    "SO2": [-520000, -100000, -50000, -10000, 0, 10000, 50000, 100000, 200000],
    "NOX": [-400000, -100000, -50000, -10000, 0, 10000, 50000, 100000, 180000],
    "VOC": [-660000, -100000, -50000, -10000, 0, 10000, 50000, 100000, 260000],
    "NH3":   [-180000, -10000, -1000, 0, 1000, 10000, 30000],
    "PM2.5": [-310000, -50000, -5000, 0, 5000, 50000, 110000],
    "PM10": [-380000, -50000, -20000, 5000, 0, 5000, 20000, 50000, 140000],
    "BC":    [-80000, -10000, -1000, 0, 1000, 10000, 30000],
    "OC": [-160000, -30000, -10000, -1000, 0, 1000, 10000, 30000, 60000]
}

pollutant_ticklabels = {
    "SO2": ['-520K', '-100K', '-50k', '-10K', '0', '+10K', '+50K', '+100K', '+200K'],
    "NOX": ['-400K', '-100K', '-50k', '-10K', '0', '+10K', '+50K', '+100K', '+180K'],
    "VOC": ['-660K', '-100K', '-50k', '-10K', '0', '+10K', '+50K', '+100K', '+260K'],
    "NH3":   ['-180K', '-10K', '-1K', '0', '+1K', '+10K', '+30K'],
    "PM2.5": ['-310K', '-50K', '-5K', '0', '+5K', '+50K', '+110K'],
    "PM10": ['-380K', '-50K', '-20k', '-5K', '0', '+5K', '+20K', '+50K', '+140K'],
    "BC":    ['-80K', '-10K', '-1K', '0', '+1K', '+10K', '+30K'],
    "OC": ['-160K', '-30k', '-10K', '-1K', '0', '+1K', '+10K', '+30K', '+60K']
}

gdf = get_province_gdf()

for pollutant in ["SO2", "NOX", "VOC", "NH3", "PM2.5", "PM10", "BC", "OC"]:
    cmap = cm.RdYlGn
    norm = pollutant_norms[pollutant]
    
    fig, ax = plt.subplots(nrows = 5, ncols = 5, figsize = (30, 25), subplot_kw = {'projection': ccrs.PlateCarree()})
    
    for index, sector in enumerate(["Power", "Industry", "Transportation", "Residential", "Agriculture"]):
        for index_2, workbook in enumerate(workbook_arr):
            sheet = workbook[pollutant]
            
            def get_color(name):
                if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
                    english_name = PROVINCE_MAP_EMISSIONS[name]
                    province_index = PROVINCE_ARR.index(english_name)
                    difference = sheet.cell(index + 12, province_index + 3).value - sheet.cell(index + 3, province_index + 3).value
                    return cmap(norm(difference))
                       
                return 'lightgrey'   
            
            total_2020 = sum(sheet.cell(index + 3, c).value for c in range(3, 34))
            total_2035 = sum(sheet.cell(index + 12, c).value for c in range(3, 34))
            
            gdf['color'] = gdf['province'].apply(get_color)
            gdf.plot(color = gdf['color'], ax = ax[index][index_2], edgecolor = 'black', linewidth = 0.1)
            
            ax[index][index_2].text(0.5, 0.02, f"2020: {total_2020/1000:.0f}K  |  2035: {total_2035/1000:.0f}K", 
                                    transform = ax[index][index_2].transAxes, fontsize = 12, ha = 'center', va = 'bottom', 
                                    bbox = dict(boxstyle = 'round,pad=0.2', fc = 'white', alpha = 0.7))
            
    plt.subplots_adjust(hspace = 0.1, wspace = 0.05)
    
    sm = cm.ScalarMappable(cmap = cmap, norm = norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax = ax, orientation = 'vertical', pad = 0.02)
    cbar.set_label('Emission Change 2020-2035 (tons)', fontsize = 14)
    cbar.set_ticks(pollutant_ticks[pollutant])
    cbar.set_ticklabels(pollutant_ticklabels[pollutant])
    
    for index, scenario in enumerate(["Baseline", "CleanAir", "OTPCA", "OTPNZCA", "EPNZCA"]):
        ax[0][index].set_title(scenario)
    for index, sector in enumerate(["Power", "Industry", "Transportation", "Residential", "Agriculture"]):
        ax[index][0].text(-0.15, 0.5, sector, transform = ax[index][0].transAxes,
                      fontsize = 14, va = 'center', rotation = 'vertical')

    fig.suptitle(f'Emission Change from 2020 to 2035 for {pollutant}', fontsize = 18, fontweight = 'bold')
    plt.savefig(f"Emission Files/Pollutant Change Province Maps/{pollutant}_Map.png")
    plt.close(fig)
    
print("Finished!")