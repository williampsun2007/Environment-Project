import openpyxl

groups = [["SO2"], ["NOX"], ["NH3"], ["VOC"], ["PM2.5"],
          ["SO2", "NOX"], ["SO2", "NH3"], ["NOX", "NH3"], ["SO2", "NOX", "NH3"],
          ["SO2", "NOX", "NH3", "VOC"], ["SO2", "NOX", "NH3", "PM2.5"]]

wb = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_2035_change_percentage.xlsx")

for group in groups:
    wb_edit = openpyxl.load_workbook("Emission Files/管控方案模板.xlsx")
    ws_edit = wb_edit["管控方案"]
    ws_edit.delete_rows(3, 10000)
    
    for pollutant in group:
        ws = wb[pollutant]
        for r in range(2, 7):
            for c in range(2, 33):
                province = ws.cell(row = 1, column = c).value
                department = ws.cell(row = r, column = 1).value
                value = ws.cell(row = r, column = c).value

                if value == "#DIV/0!":
                    value = 0
                elif float(value) < 0:
                    value = 0
                
                poll_name = pollutant if pollutant != "PM2.5" else "PM"

                ws_edit.append([province, department, poll_name, value])
                
    binary_code = ""
    for pollutant in ["SO2", "NOX", "NH3", "VOC", "PM2.5"]:
        if pollutant in group:
            binary_code += "1"
        else:
            binary_code += "0"
    
    for weather_year in range(2017, 2025):
        file_name = f"EmissionReductions_EPNZCA2035_{binary_code}_{weather_year}Met"
        wb_edit.save(f"Species Combination Test Excel Files/{file_name}-2020-{weather_year}.xlsx")