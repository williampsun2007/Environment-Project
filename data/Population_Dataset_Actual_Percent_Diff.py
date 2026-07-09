import pandas as pd
import openpyxl

china_population_dataset = {
    "北京市": [21880000, 21950000, 21940000, 21920000, 21900000, 21890000, 21890000, 21840000, 21860000, 21830000],
    "天津市": [14390000, 14430000, 14100000, 13830000, 13850000, 13870000, 13730000, 13630000, 13640000, 13640000],
    "河北省": [73450000, 73750000, 74090000, 74260000, 74470000, 74640000, 74480000, 74200000, 73930000, 73780000],
    "山西省": [35190000, 35140000, 35100000, 35020000, 34970000, 34900000, 34800000, 34810000, 34660000, 34460000],
    "内蒙古自治区": [24400000, 24360000, 24330000, 24220000, 24150000, 24030000, 24000000, 24010000, 23960000, 23880000],
    "辽宁省": [43380000, 43270000, 43120000, 42910000, 42770000, 42550000, 42290000, 41970000, 41820000, 41550000],
    "吉林省": [26130000, 25670000, 25260000, 24840000, 24480000, 23990000, 23750000, 23480000, 23390000, 23170000],
    "黑龙江省": [35290000, 34630000, 33990000, 33270000, 32550000, 31710000, 31250000, 30990000, 30620000, 30290000],
    "上海市": [24580000, 24670000, 24660000, 24750000, 24810000, 24880000, 24890000, 24750000, 24870000, 24800000],
    "江苏省": [83150000, 83810000, 84230000, 84460000, 84690000, 84770000, 85050000, 85150000, 85260000, 85260000],
    "浙江省": [59850000, 60720000, 61700000, 62730000, 63750000, 64680000, 65400000, 65770000, 66270000, 66700000],
    "安徽省": [60110000, 60330000, 60570000, 60760000, 60920000, 61050000, 61130000, 61270000, 61210000, 61230000],
    "福建省": [39840000, 40160000, 40650000, 41040000, 41370000, 41610000, 41870000, 41880000, 41830000, 41930000],
    "江西省": [44850000, 44960000, 45110000, 45130000, 45160000, 45190000, 45170000, 45280000, 45150000, 45020000],
    "山东省": [98660000, 99730000, 100330000, 100770000, 101060000, 101650000, 101700000, 101630000, 101230000, 100800000],
    "河南省": [97010000, 97780000, 98290000, 98640000, 99010000, 99410000, 98830000, 98720000, 98150000, 97850000],
    "湖北省": [58500000, 58850000, 59040000, 59170000, 59270000, 57450000, 58300000, 58440000, 58380000, 58340000],
    "湖南省": [66150000, 66250000, 66330000, 66350000, 66400000, 66450000, 66220000, 66040000, 65680000, 65390000],
    "广东省": [116780000, 119080000, 121410000, 123480000, 124890000, 126240000, 126840000, 126570000, 127060000, 127800000],
    "广西壮族自治区": [48110000, 48570000, 49070000, 49470000, 49820000, 50190000, 50370000, 50470000, 50270000, 50130000],
    "海南省": [9450000, 9570000, 9720000, 9820000, 9950000, 10120000, 10200000, 10270000, 10430000, 10480000],
    "重庆市": [30700000, 31100000, 31440000, 31630000, 31880000, 32090000, 32120000, 32130000, 31910000, 31900000],
    "四川省": [81960000, 82510000, 82890000, 83210000, 83510000, 83710000, 83720000, 83740000, 83680000, 83640000],
    "贵州省": [37080000, 37580000, 38030000, 38220000, 38480000, 38580000, 38520000, 38560000, 38650000, 38600000],
    "云南省": [46630000, 46770000, 46930000, 47030000, 47140000, 47220000, 46900000, 46930000, 46730000, 46550000],
    "西藏自治区": [3300000, 3400000, 3490000, 3540000, 3610000, 3660000, 3660000, 3640000, 3650000, 3700000],
    "陕西省": [38460000, 38740000, 39040000, 39310000, 39440000, 39550000, 39540000, 39560000, 39520000, 39530000],
    "甘肃省": [25230000, 25200000, 25220000, 25150000, 25090000, 25010000, 24900000, 24920000, 24650000, 24580000],
    "青海省": [5770000, 5820000, 5860000, 5870000, 5900000, 5930000, 5940000, 5950000, 5940000, 5930000],
    "宁夏回族自治区": [6840000, 6950000, 7050000, 7100000, 7170000, 7210000, 7250000, 7280000, 7290000, 7290000],
    "新疆维吾尔自治区": [23850000, 24280000, 24800000, 25200000, 25590000, 25900000, 25890000, 25870000, 25980000, 26230000]
}

china_population_dataset['Total'] = [1383260000, 1392320000, 1400110000, 1405410000, 1410080000, 1412120000, 1412600000, 1411750000, 1409670000, 1408280000]

df_1km = pd.read_excel("Population Graph Data/1km_Resolution_Data.xlsx")
df_prov = pd.read_excel("Population Graph Data/Province_Resolution_data.xlsx")

data_cols_1km = [c for c in df_1km.columns if '_' in str(c)]

years = list(range(2015, 2025))

province_list = ['Total'] + [p for p in df_1km['Province'].dropna().tolist() if p != 'Total']

wb = openpyxl.Workbook()
sheet = wb.active

headers = ["", "SSP1", "SSP2", "SSP3", "SSP4", "SSP5"]
for fer in range(1, 6):
    for mig in range(1, 4):
        headers.append(f"Fertility {fer}, Migration {mig}")
sheet.append(headers)

for idx, province in enumerate(province_list):
    sheet.cell(row = idx + 2, column = 1).value = province
    col_idx = 2
    
    if province == "Total":
        for scenario in ["SSP1", "SSP2", "SSP3", "SSP4", "SSP5"]:
            total_dataset = 0
            total_actual = 0
            for year in years:
                cols = [c for c in data_cols_1km if c.startswith(scenario) and c.endswith(f'_{year}')]
                vals = df_1km[cols].sum().values.astype(float)
                total_dataset += sum(vals)
                total_actual += china_population_dataset['Total'][year - 2015]
            sheet.cell(row = idx + 2, column = col_idx).value = (total_dataset - total_actual) / total_actual * 100
            col_idx += 1
            
        for fertility in range(1, 6):
            for migration in range(1, 4):
                total_dataset = 0
                total_actual = 0
                
                df_scenario = df_prov[df_prov['V2'] == f"_SSPFer{fertility}_SSPMigr{migration}"]
                
                for year in years:
                    total_dataset += df_scenario[str(year)].sum()
                    total_actual += china_population_dataset['Total'][year - 2015]
                sheet.cell(row = idx + 2, column = col_idx).value = (total_dataset - total_actual) / total_actual * 100
                col_idx += 1
    else:
        for scenario in ["SSP1", "SSP2", "SSP3", "SSP4", "SSP5"]:
            total_dataset = 0
            total_actual = 0
            df_1km_prov = df_1km[df_1km['Province'] == province]
            
            for year in years:
                cols = [c for c in data_cols_1km if c.startswith(scenario) and c.endswith(f'_{year}')]
                vals = df_1km_prov[cols].sum().values.astype(float)
                total_dataset += sum(vals)
                total_actual += china_population_dataset[province][year - 2015]
            sheet.cell(row = idx + 2, column = col_idx).value = (total_dataset - total_actual) / total_actual * 100
            col_idx += 1
            
        for fertility in range(1, 6):
            for migration in range(1, 4):
                total_dataset = 0
                total_actual = 0
                df_prov_prov = df_prov[(df_prov['V2'] == f"_SSPFer{fertility}_SSPMigr{migration}") & (df_prov['V1'] == province)]
                
                for year in years:
                    total_dataset += df_prov_prov[str(year)].values[0]
                    total_actual += china_population_dataset[province][year - 2015]
                sheet.cell(row = idx + 2, column = col_idx).value = (total_dataset - total_actual) / total_actual * 100
                col_idx += 1

wb.save("Population Graph Data/Population_Actual_Percent_Diff.xlsx")
print("Finished!")