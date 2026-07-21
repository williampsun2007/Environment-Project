'''
For each pollutant, renders a 6-panel choropleth of Chinese provinces
showing % emission change vs. 2020 (2030/2035 under two scenarios,
each vs. 2017 and vs. 2020 baselines), with national totals labeled,
and saves each figure as a PNG.
'''

import cartopy.crs as ccrs
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from cnmaps import get_adm_maps
import geopandas as gpd
from matplotlib.colors import Normalize

wb_epnzca_2020 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_Emission.xlsx")
wb_epnzca_2030 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2030_Emission.xlsx")
wb_epnzca_2035 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2035_Emission.xlsx")

wb_2017_2030 = openpyxl.load_workbook("Emission Reports/2030/emission_report-epnzca-2017 to 2030.xlsx")
wb_2017_2035 = openpyxl.load_workbook("Emission Reports/2035/emission_report-epnzca-2017 to 2035.xlsx")

wb_2020_2030 = openpyxl.load_workbook("Emission Reports/2030/emission_report-epnzca-2020 to 2030.xlsx")
wb_2020_2035 = openpyxl.load_workbook("Emission Reports/2035/emission_report-epnzca-2020 to 2035.xlsx")

province_arr = ["Beijing", "Tianjin", "Hebei", "Shanxi", "Inner Mongolia", "Liaoning", "Jilin", "Heilongjiang", "Shanghai", 
                "Jiangsu", "Zhejiang", "Anhui", "Fujian", "Jiangxi", "Shandong", "Henan", "Hubei", "Hunan", "Guangdong", 
                "Guangxi", "Hainan", "Chongqing", "Sichuan", "Guizhou", "Yunnan", "Tibet", "Shaanxi", "Gansu", "Qinghai",
                "Ningxia", "Xinjiang"]

province_map = {
    "北京市": "Beijing",
    "天津市": "Tianjin",
    "河北省": "Hebei",
    "山西省": "Shanxi",
    "内蒙古自治区": "Inner Mongolia",
    "辽宁省": "Liaoning",
    "吉林省": "Jilin",
    "黑龙江省": "Heilongjiang",
    "上海市": "Shanghai",
    "江苏省": "Jiangsu",
    "浙江省": "Zhejiang",
    "安徽省": "Anhui",
    "福建省": "Fujian",
    "江西省": "Jiangxi",
    "山东省": "Shandong",
    "河南省": "Henan",
    "湖北省": "Hubei",
    "湖南省": "Hunan",
    "广东省": "Guangdong",
    "广西壮族自治区": "Guangxi",
    "海南省": "Hainan",
    "重庆市": "Chongqing",
    "四川省": "Sichuan",
    "贵州省": "Guizhou",
    "云南省": "Yunnan",
    "西藏自治区": "Tibet",
    "陕西省": "Shaanxi",
    "甘肃省": "Gansu",
    "青海省": "Qinghai",
    "宁夏回族自治区": "Ningxia",
    "新疆维吾尔自治区": "Xinjiang"
}

records = []
for province in get_adm_maps(level = '省'):
    records.append({
        'province': province['province'],
        'geometry': province['geometry']
    })

gdf = gpd.GeoDataFrame(records, crs = 'EPSG:4326')

pollutant_ticks = {
    "SO2": [-90, -60, -30, 0, 30, 60, 90],
    "NOx": [-70, -50, -20, 0, 20, 50, 70],
    "VOC": [-90, -60, -30, 0, 30, 60, 90],
    "NH3": [-50, -30, -10, 0, 10, 30, 50],
    "PM25": [-90, -60, -30, 0, 30, 60, 90]
}

pollutant_labels = {
    "SO2": ['-90%', '-60%', '-30%', '0%', '30%', '60%', '90%'],
    "NOx": ['-70%', '-50%', '-20%', '0%', '20%', '50%', '70%'],
    "VOC": ['-90%', '-60%', '-30%', '0%', '30%', '60%', '90%'],
    "NH3": ['-50%', '-30%', '-10%', '0%', '10%', '30%', '50%'],
    "PM25": ['-90%', '-60%', '-30%', '0%', '30%', '60%', '90%']
}

def get_national_total(wb, sheet_name, row_range, col_offset):
    sheet = wb[sheet_name]
    total = 0
    for province in province_arr:
        province_index = province_arr.index(province)
        for r in row_range:
            val = sheet.cell(r, province_index + col_offset).value
            if val:
                total += val
    return total

for pollutant in ["SO2", "NOx", "VOC", "NH3", "PM25"]:
    cmap = plt.get_cmap('RdYlGn')
    vmax = pollutant_ticks[pollutant][-1]
    vmin = pollutant_ticks[pollutant][0]
    norm = Normalize(vmin = vmin, vmax = vmax) 
    poll_max = -1000
    poll_min = 1000 
    
    def get_color_2017_2030(name):
        global poll_max, poll_min
        sheet = wb_2017_2030[pollutant]
        sheet_2020 = wb_epnzca_2020[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total_2030 = 0
            for r in range(9, 14):
                total_2030 += sheet.cell(r, province_index + 5).value
                
            total_2020 = 0
            for r in range(3, 8):
                total_2020 += sheet_2020.cell(r, province_index + 3).value
            poll_max = max(poll_max, (total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            poll_min = min(poll_min, (total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            return cmap(norm((total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0))
        return "lightgrey"
    
    def get_color_2017_2035(name):
        global poll_max, poll_min
        sheet = wb_2017_2035[pollutant]
        sheet_2020 = wb_epnzca_2020[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total_2035 = 0
            for r in range(9, 14):
                total_2035 += sheet.cell(r, province_index + 5).value

            total_2020 = 0
            for r in range(3, 8):
                total_2020 += sheet_2020.cell(r, province_index + 3).value
            poll_max = max(poll_max, (total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            poll_min = min(poll_min, (total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            return cmap(norm((total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0))
        return "lightgrey"

    def get_color_2020_2030(name):
        global poll_max, poll_min
        sheet = wb_2020_2030[pollutant]
        sheet_2020 = wb_epnzca_2020[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total_2030 = 0
            for r in range(9, 14):
                total_2030 += sheet.cell(r, province_index + 5).value
            
            total_2020 = 0
            for r in range(3, 8):
                total_2020 += sheet_2020.cell(r, province_index + 3).value
            poll_max = max(poll_max, (total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            poll_min = min(poll_min, (total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            return cmap(norm((total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0))
        return "lightgrey"
    
    def get_color_2020_2035(name):
        global poll_max, poll_min
        sheet = wb_2020_2035[pollutant]
        sheet_2020 = wb_epnzca_2020[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total_2035 = 0
            for r in range(9, 14):
                total_2035 += sheet.cell(r, province_index + 5).value
            
            total_2020 = 0
            for r in range(3, 8):
                total_2020 += sheet_2020.cell(r, province_index + 3).value
            poll_max = max(poll_max, (total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            poll_min = min(poll_min, (total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            return cmap(norm((total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0))
        return "lightgrey"
    
    def get_color_2030(name):
        global poll_max, poll_min
        sheet = wb_epnzca_2030[pollutant]
        sheet_2020 = wb_epnzca_2020[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total_2030 = 0
            for r in range(3, 8):
                total_2030 += sheet.cell(r, province_index + 3).value
                
            total_2020 = 0
            for r in range(3, 8):
                total_2020 += sheet_2020.cell(r, province_index + 3).value
            poll_max = max(poll_max, (total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            poll_min = min(poll_min, (total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            return cmap(norm((total_2030 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0))
        return "lightgrey"

    def get_color_2035(name):
        global poll_max, poll_min
        sheet = wb_epnzca_2035[pollutant]
        sheet_2020 = wb_epnzca_2020[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total_2035 = 0
            for r in range(3, 8):
                total_2035 += sheet.cell(r, province_index + 3).value
                
            total_2020 = 0
            for r in range(3, 8):
                total_2020 += sheet_2020.cell(r, province_index + 3).value
            poll_max = max(poll_max, (total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            poll_min = min(poll_min, (total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0)
            return cmap(norm((total_2035 - total_2020) / total_2020 * 100 if total_2020 != 0 else 0))
        return "lightgrey"
    
    fig, ax = plt.subplots(nrows = 3, ncols = 2, figsize = (30, 25), subplot_kw = {'projection': ccrs.PlateCarree()})
    
    gdf['color'] = gdf['province'].apply(get_color_2030)
    gdf.plot(color = gdf['color'], ax = ax[0][0], edgecolor = 'black', linewidth = 0.1)
    ax[0][0].set_title(f"2030 Emissions from DPEC (% change vs 2020)")
    total = get_national_total(wb_epnzca_2030, pollutant, range(3, 8), 3)
    ax[0][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[0][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2035)
    gdf.plot(color = gdf['color'], ax = ax[0][1], edgecolor = 'black', linewidth = 0.1)
    ax[0][1].set_title(f"2035 Emissions from DPEC (% change vs 2020)")
    total = get_national_total(wb_epnzca_2035, pollutant, range(3, 8), 3)
    ax[0][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[0][1].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2017_2030)
    gdf.plot(color = gdf['color'], ax = ax[1][0], edgecolor = 'black', linewidth = 0.1)
    ax[1][0].set_title(f"2030 Emissions Relative to 2017 (% change vs 2020)")
    total = get_national_total(wb_2017_2030, pollutant, range(9, 14), 5)
    ax[1][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[1][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2017_2035)
    gdf.plot(color = gdf['color'], ax = ax[1][1], edgecolor = 'black', linewidth = 0.1)
    ax[1][1].set_title(f"2035 Emissions Relative to 2017 (% change vs 2020)")
    total = get_national_total(wb_2017_2035, pollutant, range(9, 14), 5)
    ax[1][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[1][1].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2020_2030)
    gdf.plot(color = gdf['color'], ax = ax[2][0], edgecolor = 'black', linewidth = 0.1)
    ax[2][0].set_title(f"2030 Emissions Relative to 2020 (% change vs 2020)")
    total = get_national_total(wb_2020_2030, pollutant, range(9, 14), 5)
    ax[2][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[2][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2020_2035)
    gdf.plot(color = gdf['color'], ax = ax[2][1], edgecolor = 'black', linewidth = 0.1)
    ax[2][1].set_title(f"2035 Emissions Relative to 2020 (% change vs 2020)")
    total = get_national_total(wb_2020_2035, pollutant, range(9, 14), 5)
    ax[2][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[2][1].transAxes, ha = 'center', fontsize = 10)
            
    plt.subplots_adjust(hspace = 0.1, wspace = 0.05)
    
    sm = cm.ScalarMappable(cmap = cmap, norm = norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax = ax, orientation = 'vertical', pad = 0.02)
    cbar.set_label('Percent Change from 2020 Emissions (%)', fontsize = 14)
    cbar.set_ticks(pollutant_ticks[pollutant])
    cbar.set_ticklabels(pollutant_labels[pollutant])
    
    print(f"Pollutant: {pollutant}, Max: {poll_max}, Min: {poll_min}")

    fig.suptitle(f'Emission Percent Change from 2020 for {pollutant} (EPNZCA Scenario)', fontsize = 18, fontweight = 'bold')
    plt.savefig(f"C:/Users/sunyi/Environment_Project/Emission Files/Percent Change from 2020 Baseline/EPNZCA/{pollutant}_Map.png")
    plt.close(fig)
    
print("Finished!")