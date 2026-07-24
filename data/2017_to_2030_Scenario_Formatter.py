'''
Computes 2017->2030 % emission reduction per pollutant/sector/region,
then reshapes those values into the control-plan template (clamping
invalid/negative entries to 0) and saves one scenario file per
weather year (2017-2024).
'''

import openpyxl

wb_2017 = openpyxl.load_workbook("2017_Base_Data/emission_report.xlsx")
wb_2030 = openpyxl.load_workbook("EPNZCA_All_Years/early_peak-net_zero-clean_air_2030_Emission.xlsx")
wb_percentage = openpyxl.load_workbook("early_peak-net_zero-clean_air/2017_2030_change_percentage.xlsx")

for sheet_name in ["SO2", "NOx", "VOC", "NH3", "PM10", "PM25", "BC", "OC"]:
    sheet_2017 = wb_2017[sheet_name]
    sheet_2030 = wb_2030[sheet_name]
    sheet_percentage = wb_percentage[sheet_name]
    
    for r in range(3, 8):
        for c in range(3, 34):
            real_r = r - 1
            if r == 5:
                real_r = 5
            elif r == 6:
                real_r = 4
            
            value_2017 = sheet_2017.cell(row = real_r, column = c + 2).value
            value_2030 = sheet_2030.cell(row = r, column = c).value
            
            if value_2017 != 0:
                percent_difference = ((value_2017 - value_2030) / value_2017) * 100
            else:
                percent_difference = 0
                
            sheet_percentage.cell(row = r - 1, column = c - 1).value = percent_difference
            
wb_percentage.save("early_peak-net_zero-clean_air/2017_2030_change_percentage.xlsx")

wb_edit = openpyxl.load_workbook("Emission Files/管控方案模板.xlsx")
ws_edit = wb_edit["管控方案"]
ws_edit.delete_rows(3, 10000)
    
for sheet_name in ["SO2", "NOx", "VOC", "NH3", "PM25"]:
    this_sheet_name = sheet_name
    ws = wb_percentage[this_sheet_name]
    
    for r in range(2, 7):
        for c in range(2, 33):
            province = ws.cell(row = 1, column = c).value
            department = ws.cell(row = r, column = 1).value
            value = ws.cell(row = r, column = c).value

            if value == "#DIV/0!":
                value = 0
            elif float(value) < 0:
                value = 0
                
            if sheet_name == "PM25":
                sheet_name = "PM"

            ws_edit.append([province, department, sheet_name, value])
            
for weather_year in range(2017, 2025):
    wb_edit.save(f"Scenario Excel Files 2017-2030/early-peak-net-zero-clean-air-2017-{weather_year}_2017_{weather_year}.xlsx")

print("Finished!")