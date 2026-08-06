from pathlib import Path
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

df_co2 = pd.read_excel("Emission Files/Historical Emission/CO2 Emissions.xlsx", sheet_name = "CO2")
df_co2 = df_co2[df_co2['Unnamed: 0'] >= 2010]
df_co2 = df_co2.set_index("Unnamed: 0")

df_historical_v2_co2 = pd.read_excel("Emission Files/Historical Emission/ABaCAS-EI v2.0 dataset.xlsx", sheet_name = "CO2")
df_historical_v2_co2 = df_historical_v2_co2[df_historical_v2_co2['Year'] >= 2010]
df_historical_v2_co2 = df_historical_v2_co2[['Year', 'China']]
df_historical_v2_co2 = df_historical_v2_co2.groupby('Year').sum().reset_index().drop('Year', axis = 1) * 1000

df_historical_v2_pm25 = pd.read_excel("Emission Files/Historical Emission/ABaCAS-EI v2.0 dataset.xlsx", sheet_name = "PM2.5")
df_historical_v2_pm25 = df_historical_v2_pm25[df_historical_v2_pm25['Year'] >= 2010]
df_historical_v2_pm25 = df_historical_v2_pm25[['Year', 'China']]
df_historical_v2_pm25 = df_historical_v2_pm25.groupby('Year').sum().reset_index().drop('Year', axis = 1) * 1000

df_historical_v2_so2 = pd.read_excel("Emission Files/Historical Emission/ABaCAS-EI v2.0 dataset.xlsx", sheet_name = "SO2")
df_historical_v2_so2 = df_historical_v2_so2[df_historical_v2_so2['Year'] >= 2010]
df_historical_v2_so2 = df_historical_v2_so2[['Year', 'China']]
df_historical_v2_so2 = df_historical_v2_so2.groupby('Year').sum().reset_index().drop('Year', axis = 1) * 1000

df_historical_v2_nox = pd.read_excel("Emission Files/Historical Emission/ABaCAS-EI v2.0 dataset.xlsx", sheet_name = "NOx")
df_historical_v2_nox = df_historical_v2_nox[df_historical_v2_nox['Year'] >= 2010]
df_historical_v2_nox = df_historical_v2_nox[['Year', 'China']]
df_historical_v2_nox = df_historical_v2_nox.groupby('Year').sum().reset_index().drop('Year', axis = 1) * 1000

df_historical_v2_nh3 = pd.read_excel("Emission Files/Historical Emission/ABaCAS-EI v2.0 dataset.xlsx", sheet_name = "NH3")
df_historical_v2_nh3 = df_historical_v2_nh3[df_historical_v2_nh3['Year'] >= 2010]
df_historical_v2_nh3 = df_historical_v2_nh3[['Year', 'China']]
df_historical_v2_nh3 = df_historical_v2_nh3.groupby('Year').sum().reset_index().drop('Year', axis = 1) * 1000

df_historical_v2_voc = pd.read_excel("Emission Files/Historical Emission/ABaCAS-EI v2.0 dataset.xlsx", sheet_name = "VOCs")
df_historical_v2_voc = df_historical_v2_voc[df_historical_v2_voc['Year'] >= 2010]
df_historical_v2_voc = df_historical_v2_voc[['Year', 'China']]
df_historical_v2_voc = df_historical_v2_voc.groupby('Year').sum().reset_index().drop('Year', axis = 1) * 1000

scenarios = ["Baseline", "CleanAir", "OTPCA", "OTPNZCA", "EPNZCA"]
pollutants = ["CO2", "SO2", "NOx", "NH3", "VOC", "PM25"]
file_prefix = {
    "Baseline": "baseline",
    "CleanAir": "clean_air",
    "OTPCA": "on-time_peak-clean_air",
    "OTPNZCA": "on-time_peak-net_zero-clean_air",
    "EPNZCA": "early_peak-net_zero-clean_air",
}

scenario_dict = {}
for scenario in scenarios:
    scenario_dict[scenario] = {p: [] for p in pollutants}
    folder_path = Path(f"{scenario}_All_Years")

    for year in [2020, 2025, 2030]:
        file = folder_path / f"{file_prefix[scenario]}_{year}_Emission.xlsx"
        workbook = openpyxl.load_workbook(file)

        for pollutant in pollutants:
            sheet = workbook[pollutant]
            total = sum(sheet.cell(row = 2, column = c).value for c in range(3, 34))
            scenario_dict[scenario][pollutant].append(total)

fig, ax = plt.subplots(figsize = (14, 8), nrows = 2, ncols = 3)
axes = ax.flatten()

for i, pollutant in enumerate(pollutants):
    for scenario in scenarios:
        axes[i].plot([2020, 2025, 2030], scenario_dict[scenario][pollutant], label = scenario)
        
    if pollutant != "CO2":
        historical_emission = []
        for year in range(2010, 2024):
            wb = openpyxl.load_workbook(f"Emission Files/Historical Emission/{year} Emissions.xlsx")
            sheet = wb[pollutant]
            total = sum(float(sheet.cell(row = r, column = c).value) for c in range(4, 35) for r in range(2, 7) if sheet.cell(row = r, column = c).value)
            historical_emission.append(total)
        axes[i].scatter(list(range(2010, 2024)), historical_emission, label = "Historical", marker = "o", color = "black")
        
        if pollutant == "SO2":
            axes[i].scatter(list(range(2010, 2022)), df_historical_v2_so2, label = "Historical v2", marker = "o", color = "purple")
        elif pollutant == "NOx":
            axes[i].scatter(list(range(2010, 2022)), df_historical_v2_nox, label = "Historical v2", marker = "o", color = "purple")
        elif pollutant == "NH3":
            axes[i].scatter(list(range(2010, 2022)), df_historical_v2_nh3, label = "Historical v2", marker = "o", color = "purple")
        elif pollutant == "VOC":
            axes[i].scatter(list(range(2010, 2022)), df_historical_v2_voc, label = "Historical v2", marker = "o", color = "purple")
        else:
            axes[i].scatter(list(range(2010, 2022)), df_historical_v2_pm25, label = "Historical v2", marker = "o", color = "purple")
    else:
        df_co2_china = df_co2['China'] * (44 / 12) * 1000000
        axes[i].scatter(list(range(2010, 2025)), df_co2_china, label = "Historical", marker = "o", color = "black")
        axes[i].scatter(list(range(2010, 2022)), df_historical_v2_co2, label = "Historical v2", marker = "o", color = "purple")
    
    axes[i].set_xticks(list(range(2010, 2031, 4)))
    axes[i].set_ylabel("Emission (tons)")
    axes[i].set_title(pollutant)
    axes[i].legend(fontsize = 6)

fig.suptitle("Emission Comparison for Different Scenarios", fontsize = 16)
plt.tight_layout()
plt.savefig("Emission Files/Scenario Emission Graphs/DPEC_Scenario_Comparison_2010.png")
plt.show()