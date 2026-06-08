import openpyxl

wb_2020 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\baseline\baseline_2020_Emission.xlsx")
wb_2035 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\baseline\clean_air_2035_Emission.xlsx")

for sheet_name in ['SO2', 'NOx', 'VOC', 'NH3', 'PM25', 'PM10', 'BC', 'OC']:
    total_species_reduction = 0
    total_species_increase = 0
    
    ws_2020 = wb_2020[sheet_name]
    ws_2035 = wb_2035[sheet_name]
    
    for r in range(3, 8):
        for c in range(3, 34):
            value_2020 = ws_2020.cell(row = r, column = c).value
            value_2035 = ws_2035.cell(row = r, column = c).value
            
            total_species_reduction += (value_2020 - value_2035)
            
            if (value_2035 > value_2020):
                total_species_increase += (value_2035 - value_2020)
          
    print(f"{sheet_name} Total Reduction: {total_species_reduction}")
    print(f"{sheet_name} Total Increase: {total_species_increase}")
    print(f"Increase relative to Reduction: {total_species_increase / total_species_reduction * 100:.2f}%")
    print("--------------------------------------------------")
            