'''
For each pollutant and reduction percent (10-100%), builds a
control-plan workbook applying that flat reduction uniformly across
all province/department rows, and saves one file per weather year
(2017-2024).
'''

import openpyxl

wb = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_2035_change_percentage.xlsx")
ws = wb["SO2"]

for sheet_name in ["SO2", "NOX", "VOC", "NH3", "PM2.5"]:
    for p in range(10, 101, 10):
        wb_edit = openpyxl.load_workbook("Emission Files/管控方案模板.xlsx")
        ws_edit = wb_edit["管控方案"]
        ws_edit.delete_rows(3, 10000)
        for r in range(2, 7):
            for c in range(2, 33):
                province = ws.cell(row = 1, column = c).value
                department = ws.cell(row = r, column = 1).value
                
                breeze_name = "PM" if sheet_name == "PM2.5" else sheet_name
                ws_edit.append([province, department, breeze_name, p])

        for weather_year in range(2017, 2025):
            wb_edit.save(f"Scenario Excel Files Percentage Decrease/EmissionReductions_{breeze_name}_{p}pct_{weather_year}_2020Base-2020-{weather_year}.xlsx")

print("Finished!")
