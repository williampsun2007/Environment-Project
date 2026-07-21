'''
Extracts PM25.nc from each downloaded scenario zip, computes the %
of Chinese grid cells with PM2.5 >= 25 ug/m3 for the annual mean and
every day of the year, and saves one sheet per scenario/met-year to
a workbook.
'''

import xarray as xr
from pathlib import Path
import zipfile
import openpyxl
from common import get_china_geometry, build_in_china_mask

wb = openpyxl.Workbook()
wb.remove(wb.active)

file_path = Path("Excel Downloads")

list_zips = file_path.glob("*.zip")

china_geom = get_china_geometry()

abbrev = {
    "baseline": "base",
    "clean-air": "ca",
    "early-peak-net-zero-clean-air": "epnzca",
    "on-time-peak-clean-air": "otpca",
    "on-time-peak-net-zero-clean-air": "otpnzca"
}

count = 1
for file_path in list_zips:
    with zipfile.ZipFile(file_path, "r") as z:
        with z.open("PM25.nc") as src:
            base_name = Path(file_path).stem.split("-2020-")
            zip_name = abbrev.get(base_name[0])
            wb.create_sheet(title = f"{zip_name}-{base_name[1]}")
            
            year = base_name[1].split("_")[0]
            output_path = Path(file_path).parent.parent / "NC Files and Emission Reports" / f"EmissionReductions2035_{base_name[0]}_{year}Met.nc"
            with open(output_path, "wb") as dst:
                dst.write(src.read())
                
    ds = xr.open_dataset(output_path)
    
    in_china = build_in_china_mask(ds["lon"].values, ds["lat"].values, china_geom=china_geom)
    
    day_data = ds["pred_PM25"].mean(dim="time").values
    china_cells = day_data[in_china]
    pct = (china_cells >= 25).sum() / len(china_cells) * 100
    
    wb[f"{zip_name}-{base_name[1]}"].append(["Annual Mean", pct])
    
    for day in range(365):
        day_data = ds["pred_PM25"][day].values
        china_cells = day_data[in_china]
        pct = (china_cells >= 25).sum() / len(china_cells) * 100
        wb[f"{zip_name}-{base_name[1]}"].append([f"Day {day + 1}", pct])
    
        print(f"{count} {zip_name}-{base_name[1]} Day {day + 1}: {pct:.2f}% of China cells >= 25")
    
    count += 1
    
print(wb.sheetnames)
    
wb.save("Emission Files/Percent_China_Over_25.xlsx")
print("Done!")