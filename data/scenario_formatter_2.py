import openpyxl

wb_2020 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_Emission.xlsx")
wb_2030 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2030_Emission.xlsx")
wb_percentage = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_2030_change_percentage.xlsx")

for sheet_name in ["SO2", "NOx", "VOC", "NH3", "PM10", "PM25", "BC", "OC"]:
    sheet_2020 = wb_2020[sheet_name]
    sheet_2030 = wb_2030[sheet_name]
    sheet_percentage = wb_percentage[sheet_name]
    
    for r in range(3, 8):
        for c in range(3, 34):
            value_2020 = sheet_2020.cell(row = r, column = c).value
            value_2030 = sheet_2030.cell(row = r, column = c).value
            
            if value_2020 != 0:
                percent_difference = ((value_2020 - value_2030) / value_2020) * 100
            else:
                percent_difference = 0
                
            sheet_percentage.cell(row = r - 1, column = c - 1).value = percent_difference
            
wb_percentage.save("early_peak-net_zero-clean_air/2020_2030_change_percentage.xlsx")

wb_edit = openpyxl.load_workbook("Emission Files/管控方案模板.xlsx")
ws_edit = wb_edit["管控方案"]
ws_edit.delete_rows(3, 10000)
    
for sheet_name in ["SO2", "NOx", "VOC", "NH3", "PM25"]:
    this_sheet_name = sheet_name
    if this_sheet_name == "NOX":
        this_sheet_name = "NOx"
    ws = wb_percentage[this_sheet_name]
    
    for r in range(2, 7):
        for c in range(2, 33):
            province = ws.cell(row = 1, column = c).value
            department = ws.cell(row = r, column = 1).value
            value = ws.cell(row = r, column = c).value

            if value == "#DIV/0!":
                value = 0
            elif float(value) < 0:
                value = 0
                
            if sheet_name == "PM25":
                sheet_name = "PM"

            ws_edit.append([province, department, sheet_name, value])
            
for weather_year in range(2017, 2025):
    wb_edit.save(f"Scenario Excel Files 2020-2030/early_peak-net_zero-clean_air-2020-{weather_year}_2020_{weather_year}.xlsx")

print("Finished!")