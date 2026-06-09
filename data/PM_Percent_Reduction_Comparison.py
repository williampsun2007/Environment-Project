import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_2035_change_percentage.xlsx")
wb_raw = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_2035_Raw.xlsx")

region_sector_reduction = {}

red_font = Font(color="FF0000", bold=True)  # Uses hex color code

for sheet_name in ['PM2.5', 'PM10', 'BC', 'OC']:
    ws = wb[sheet_name]
    for r in range(2, 7):
        for c in range(2, 33):
            province = ws.cell(row = 1, column = c).value
            department = ws.cell(row = r, column = 1).value
            value = ws.cell(row = r, column = c).value
            
            if value == "#DIV/0!":
                value = 0
                
            value = round(value)

            if region_sector_reduction.get(f"{province}_{department}") is None:
                region_sector_reduction[f"{province}_{department}"] = []
            
            region_sector_reduction[f"{province}_{department}"].append(value)
            
for sheet_name in ['PM2.5', 'PM10', 'BC', 'OC']:
    ws = wb_raw[sheet_name]
    for r in range(2, 7):
        for c in range(2, 33):
            province = ws.cell(row = 20, column = c + 1).value
            department = ws.cell(row = r + 19, column = 2).value
            value_2020 = ws.cell(row = r + 1, column = c + 1).value
            value_2035 = ws.cell(row = r + 10, column = c + 1).value
            
            if value_2020 is None:
                value_2020 = 0
            if value_2035 is None:
                value_2035 = 0
            
            region_sector_reduction[f"{province}_{department}"].extend([round(value_2020, 2), round(value_2035, 2)])
 
 
wb_edit = openpyxl.Workbook()

wb_edit.create_sheet(title = "PM_Reduction_Comparison")
wb_edit["PM_Reduction_Comparison"].append(['Province', 'Department', 'PM2.5_Reduction', 'PM10_Reduction', 'BC_Reduction', 'OC_Reduction',
                             'PM2.5_2020', 'PM2.5_2035', 'PM10_2020', 'PM10_2035', 'BC_2020', 'BC_2035', 'OC_2020', 'OC_2035'])

index = 0           
for key, value in region_sector_reduction.items():
    parts = key.split("_")
    province = parts[0]
    department = parts[1]
    
    max_pm = max(value[0], value[1])
    min_pm = min(value[0], value[1])
    max_oc_ec = max(value[2], value[3])
    min_oc_ec = min(value[2], value[3])
    highlight_red = (max_pm > 0 and min_oc_ec < 0) or (max_oc_ec > 0 and min_pm < 0)
    
    wb_edit["PM_Reduction_Comparison"].append([province, department, value[0], value[1], value[2], value[3], value[4], value[5], value[6], value[7], 
                                value[8], value[9], value[10], value[11]])
    
    if highlight_red:
        new_row_idx = wb_edit["PM_Reduction_Comparison"].max_row
        for cell in wb_edit["PM_Reduction_Comparison"][new_row_idx]:
            cell.font = red_font
    
    index += 1

wb_edit.save("early_peak-net_zero-clean_air/PM_Reduction_Comparison.xlsx")

print("Finished!")
    
    