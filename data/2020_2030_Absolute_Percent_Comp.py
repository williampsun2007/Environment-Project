import cartopy.crs as ccrs
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from cnmaps import get_adm_maps
import geopandas as gpd
from matplotlib.colors import Normalize

wb_baseline_2030 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\baseline\baseline_2030_Emission.xlsx")
wb_baseline_2035 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\baseline\baseline_2035_Emission.xlsx")

wb_2017_2030 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\Emission Files\emission_report - 2017 to 2030.xlsx")
wb_2017_2035 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\Emission Files\emission_report - 2017 to 2035.xlsx")

wb_2020_2030 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\Emission Files\emission_report - 2020 to 2030.xlsx")
wb_2020_2035 = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\Emission Files\emission_report - 2020 to 2035.xlsx")

province_arr = ["Beijing", "Tianjin", "Hebei", "Shanxi", "Inner Mongolia", "Liaoning", "Jilin", "Heilongjiang", "Shanghai", 
                "Jiangsu", "Zhejiang", "Anhui", "Fujian", "Jiangxi", "Shandong", "Henan", "Hubei", "Hunan", "Guangdong", 
                "Guangxi", "Hainan", "Chongqing", "Sichuan", "Guizhou", "Yunnan", "Tibet", "Shaanxi", "Gansu", "Qinghai",
                "Ningxia", "Xinjiang"]

province_map = {
    "北京市": "Beijing",
    "天津市": "Tianjin",
    "河北省": "Hebei",
    "山西省": "Shanxi",
    "内蒙古自治区": "Inner Mongolia",
    "辽宁省": "Liaoning",
    "吉林省": "Jilin",
    "黑龙江省": "Heilongjiang",
    "上海市": "Shanghai",
    "江苏省": "Jiangsu",
    "浙江省": "Zhejiang",
    "安徽省": "Anhui",
    "福建省": "Fujian",
    "江西省": "Jiangxi",
    "山东省": "Shandong",
    "河南省": "Henan",
    "湖北省": "Hubei",
    "湖南省": "Hunan",
    "广东省": "Guangdong",
    "广西壮族自治区": "Guangxi",
    "海南省": "Hainan",
    "重庆市": "Chongqing",
    "四川省": "Sichuan",
    "贵州省": "Guizhou",
    "云南省": "Yunnan",
    "西藏自治区": "Tibet",
    "陕西省": "Shaanxi",
    "甘肃省": "Gansu",
    "青海省": "Qinghai",
    "宁夏回族自治区": "Ningxia",
    "新疆维吾尔自治区": "Xinjiang"
}

records = []
for province in get_adm_maps(level = '省'):
    records.append({
        'province': province['province'],
        'geometry': province['geometry']
    })

gdf = gpd.GeoDataFrame(records, crs = 'EPSG:4326')

pollutant_max = {
    'SO2': 1_000_000,
    'NOx': 2_500_000,
    'VOC': 3_000_000,
    'NH3': 1_000_000,
    'PM25': 750_000
}

pollutant_ticks = {
    "SO2": [0, 250000, 500000, 1000000],
    "NOx": [0, 1000000, 2000000, 2500000],
    "VOC": [0, 1000000, 2000000, 3000000],
    "NH3": [0, 400000, 800000, 1000000],
    "PM25": [0, 300000, 500000, 750000]
}

pollutant_ticklabels = {
    "SO2": ["0", "250k", "500k", "1M"],
    "NOx": ["0", "1M", "2M", "2.5M"],
    "VOC": ["0", "1M", "2M", "3M"],
    "NH3": ["0", "400k", "800k", "1M"],
    "PM25": ["0", "300k", "500k", "750k"]
}

def get_national_total(wb, sheet_name, row_range, col_offset):
    sheet = wb[sheet_name]
    total = 0
    for province in province_arr:
        province_index = province_arr.index(province)
        for r in row_range:
            val = sheet.cell(r, province_index + col_offset).value
            if val:
                total += val
    return total

for pollutant in ["SO2", "NOx", "VOC", "NH3", "PM25"]:
    cmap = plt.cm.YlOrRd
    norm = Normalize(vmin = 0, vmax = pollutant_max[pollutant])    
    
    def get_color_2017_2030(name):
        sheet = wb_2017_2030[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"
    
    def get_color_2017_2035(name):
        sheet = wb_2017_2035[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"

    def get_color_2020_2030(name):
        sheet = wb_2020_2030[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"
    
    def get_color_2020_2035(name):
        sheet = wb_2020_2035[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"
    
    def get_color_2030(name):
        sheet = wb_baseline_2030[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total = 0
            for r in range(3, 8):
                total += sheet.cell(r, province_index + 3).value
            return cmap(norm(total))
        return "lightgrey"

    def get_color_2035(name):
        sheet = wb_baseline_2035[pollutant]
        if name in province_map and name not in ["台湾省", "香港特别行政区", "澳门特别行政区"]:
            english_name = province_map[name]
            province_index = province_arr.index(english_name)
            total = 0
            for r in range(3, 8):
                total += sheet.cell(r, province_index + 3).value
            return cmap(norm(total))
        return "lightgrey"
    
    fig, ax = plt.subplots(nrows = 3, ncols = 2, figsize = (30, 25), subplot_kw = {'projection': ccrs.PlateCarree()})
    
    gdf['color'] = gdf['province'].apply(get_color_2030)
    gdf.plot(color = gdf['color'], ax = ax[0][0], edgecolor = 'black', linewidth = 0.1)
    ax[0][0].set_title("2030 Emissions from DPEC")
    total = get_national_total(wb_baseline_2030, pollutant, range(3, 8), 3)
    ax[0][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[0][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2035)
    gdf.plot(color = gdf['color'], ax = ax[0][1], edgecolor = 'black', linewidth = 0.1)
    ax[0][1].set_title("2035 Emissions from DPEC")
    total = get_national_total(wb_baseline_2035, pollutant, range(3, 8), 3)
    ax[0][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[0][1].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2017_2030)
    gdf.plot(color = gdf['color'], ax = ax[1][0], edgecolor = 'black', linewidth = 0.1)
    ax[1][0].set_title("2030 Emissions Relative to 2017")
    total = get_national_total(wb_2017_2030, pollutant, range(9, 14), 5)
    ax[1][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[1][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2017_2035)
    gdf.plot(color = gdf['color'], ax = ax[1][1], edgecolor = 'black', linewidth = 0.1)
    ax[1][1].set_title("2035 Emissions Relative to 2017")
    total = get_national_total(wb_2017_2035, pollutant, range(9, 14), 5)
    ax[1][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[1][1].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2020_2030)
    gdf.plot(color = gdf['color'], ax = ax[2][0], edgecolor = 'black', linewidth = 0.1)
    ax[2][0].set_title("2030 Emissions Relative to 2020")
    total = get_national_total(wb_2020_2030, pollutant, range(9, 14), 5)
    ax[2][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[2][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2020_2035)
    gdf.plot(color = gdf['color'], ax = ax[2][1], edgecolor = 'black', linewidth = 0.1)
    ax[2][1].set_title("2035 Emissions Relative to 2020")
    total = get_national_total(wb_2020_2035, pollutant, range(9, 14), 5)
    ax[2][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[2][1].transAxes, ha = 'center', fontsize = 10)
            
    plt.subplots_adjust(hspace = 0.1, wspace = 0.05)
    
    sm = cm.ScalarMappable(cmap = cmap, norm = norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax = ax, orientation = 'vertical', pad = 0.02)
    cbar.set_label('Absolute Emission Amount (tons)', fontsize = 14)
    cbar.set_ticks(pollutant_ticks[pollutant])
    cbar.set_ticklabels(pollutant_ticklabels[pollutant])

    fig.suptitle(f'Emission Amounts for {pollutant}', fontsize = 18, fontweight = 'bold')
    plt.savefig(f"C:/Users/sunyi/Environment_Project/Emission Files/Absolute Emissions Per Province/{pollutant}_Map.png")
    plt.close(fig)
    
print("Finished!")