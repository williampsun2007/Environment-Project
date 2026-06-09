import openpyxl

wb_2020 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_Emission.xlsx")
wb_2035 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2035_Emission.xlsx")
wb = openpyxl.load_workbook("Emission Files/Total Reduction and Increase Per Pollutant.xlsx")

for sheet_name in ['SO2', 'NOx', 'VOC', 'NH3', 'PM25', 'PM10', 'BC', 'OC']:
    total_species_reduction = 0
    total_species_increase = 0
    
    ws_2020 = wb_2020[sheet_name]
    ws_2035 = wb_2035[sheet_name]
    
    for r in range(3, 8):
        for c in range(3, 34):
            value_2020 = ws_2020.cell(row = r, column = c).value
            value_2035 = ws_2035.cell(row = r, column = c).value
            
            total_species_reduction += (value_2020 - value_2035)
            
            if (value_2035 > value_2020):   
                total_species_increase += (value_2035 - value_2020)
                
    wb[sheet_name].append(["early_peak-net_zero-clean_air", total_species_reduction, total_species_increase, 
                           total_species_increase / total_species_reduction * 100 if total_species_reduction != 0 else 0])
    
wb.save("Emission Files/Total Reduction and Increase Per Pollutant.xlsx")
    
print("Finished!")
            