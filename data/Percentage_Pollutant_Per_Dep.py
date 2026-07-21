'''
For each pollutant sheet, converts raw sector emissions (Power,
Industry, Transportation, Residential, Agriculture) by province into
% share of total for 2020 and 2035, highlighting shares >= 50% in
red, and saves the reformatted workbook.
'''

import openpyxl
from openpyxl.styles import Font

wb_raw = openpyxl.load_workbook("on-time_peak-net_zero-clean_air/2020_2035_Raw.xlsx")

wb_edit = openpyxl.Workbook()

provinces = ["北京市", "天津市", "河北省", "山西省", "内蒙古自治区", "辽宁省", "吉林省", 
             "黑龙江省", "上海市", "江苏省", "浙江省", "安徽省", "福建省", "江西省", 
             "山东省", "河南省", "湖北省", "湖南省", "广东省", "广西壮族自治区", "海南省", 
             "重庆市", "四川省", "贵州省", "云南省", "西藏自治区", "陕西省", "甘肃省", 
             "青海省", "宁夏回族自治区", "新疆维吾尔自治区"]

red_font = Font(color = "FF0000", bold = True)

for sheet_name in wb_raw.sheetnames:
    wb_edit.create_sheet(title = sheet_name)
    wb_edit[sheet_name].cell(row = 1, column = 1).value = "2020"
    wb_edit[sheet_name].append([None] + provinces)
    wb_edit[sheet_name].cell(row = 3, column = 1).value = "电力"
    wb_edit[sheet_name].cell(row = 4, column = 1).value = "工业"
    wb_edit[sheet_name].cell(row = 5, column = 1).value = "交通"
    wb_edit[sheet_name].cell(row = 6, column = 1).value = "民用"
    wb_edit[sheet_name].cell(row = 7, column = 1).value = "农业"
    
    for c in range(3, 34):
        this_total = 0
        for r in range(3, 8):
            this_total += wb_raw[sheet_name].cell(row = r, column = c).value
        
        for r in range(3, 8):
            value = wb_raw[sheet_name].cell(row = r, column = c).value
            wb_edit[sheet_name].cell(row = r, column = c - 1).value = round(value / this_total * 100, 2) if this_total != 0 else 0
            if wb_edit[sheet_name].cell(row = r, column = c - 1).value >= 50:
                wb_edit[sheet_name].cell(row = r, column = c - 1).font = red_font
                
    wb_edit[sheet_name].cell(row = 10, column = 1).value = "2035"
    wb_edit[sheet_name].append([None] + provinces)
    wb_edit[sheet_name].cell(row = 12, column = 1).value = "电力"
    wb_edit[sheet_name].cell(row = 13, column = 1).value = "工业"
    wb_edit[sheet_name].cell(row = 14, column = 1).value = "交通"
    wb_edit[sheet_name].cell(row = 15, column = 1).value = "民用"
    wb_edit[sheet_name].cell(row = 16, column = 1).value = "农业"
    
    for c in range(3, 34):
        this_total = 0
        for r in range(12, 17):
            this_total += wb_raw[sheet_name].cell(row = r, column = c).value
        
        for r in range(12, 17):
            value = wb_raw[sheet_name].cell(row = r, column = c).value
            wb_edit[sheet_name].cell(row = r, column = c - 1).value = round(value / this_total * 100, 2) if this_total != 0 else 0
            if wb_edit[sheet_name].cell(row = r, column = c - 1).value >= 50:
                wb_edit[sheet_name].cell(row = r, column = c - 1).font = red_font
                
wb_edit.save("on-time_peak-net_zero-clean_air/Percentage_Pollutant_Per_Dep.xlsx")
print("Finished!")