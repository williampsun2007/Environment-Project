'''
For each province+pollutant pair, applies an extra 10% cut to that
one province's 2035 value on top of the baseline scenario, recomputes
% reduction from 2020 for every cell, and saves one control-plan
workbook per province/pollutant combination.
'''

import openpyxl

wb_2020 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2020_Emission.xlsx")
wb_2035 = openpyxl.load_workbook("early_peak-net_zero-clean_air/2035_Emission.xlsx")

provinces = ["北京市", "天津市", "河北省", "山西省", "内蒙古自治区", "辽宁省", "吉林省", 
             "黑龙江省", "上海市", "江苏省", "浙江省", "安徽省", "福建省", "江西省", 
             "山东省", "河南省", "湖北省", "湖南省", "广东省", "广西壮族自治区", "海南省", 
             "重庆市", "四川省", "贵州省", "云南省", "西藏自治区", "陕西省", "甘肃省", 
             "青海省", "宁夏回族自治区", "新疆维吾尔自治区"]

sectors = ["电力", "工业", "交通", "民用", "农业"]

pollutants = ["SO2", "NOx", "NH3", "VOC", "PM25"]

for index, province in enumerate(provinces, 1):
    for index_2, pollutant in enumerate(pollutants, 1):
        wb_edit = openpyxl.load_workbook("Emission Files/管控方案模板.xlsx")
        ws_edit = wb_edit["管控方案"]
        ws_edit.delete_rows(3, 10000)
    
        for this_pollutant in ["SO2", "NOx", "NH3", "VOC", "PM25"]:
            ws_2020 = wb_2020[this_pollutant]
            ws_2035 = wb_2035[this_pollutant]
            for r in range(3, 8):
                for c in range(3, 34):
                    this_province = provinces[c - 3]
                    this_sector = sectors[r - 3]
                    value_2020 = ws_2020.cell(row = r, column = c).value
                    value_2035 = ws_2035.cell(row = r, column = c).value
                    
                    if this_province == province and this_pollutant == pollutant:
                        value_2035 = value_2035 * 0.9
                        
                    pct_decrease = (value_2020 - value_2035) / value_2020 * 100 if value_2020 != 0 else 0
                    if pct_decrease < 0:
                        pct_decrease = 0
                
                    poll_name = this_pollutant if this_pollutant != "PM25" else "PM"

                    ws_edit.append([this_province, this_sector, poll_name, pct_decrease])

        file_name = f"EmissionReductions_EPNZCA2035_OP1_PR{index}_SP{index_2}_10pctRed_2020Met"
        wb_edit.save(f"EPNZCA Extra Percent Reduction Excel Files/{file_name}-2020-{2020}.xlsx")