'''
For each pollutant, renders a 6-panel choropleth of Chinese provinces
showing absolute emission totals (baseline 2030/2035, plus 2030/2035
under two other scenarios vs. 2017 and vs. 2020), with national
totals labeled, and saves each figure as a PNG.
'''

import cartopy.crs as ccrs
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from matplotlib.colors import Normalize
from common import PROVINCE_MAP_EMISSIONS, PROVINCE_ARR, EXCLUDED_REGIONS, get_national_total, get_province_gdf

wb_baseline_2030 = openpyxl.load_workbook("baseline/baseline_2030_Emission.xlsx")
wb_baseline_2035 = openpyxl.load_workbook("baseline/baseline_2035_Emission.xlsx")

wb_2017_2030 = openpyxl.load_workbook("Emission Files/emission_report - 2017 to 2030.xlsx")
wb_2017_2035 = openpyxl.load_workbook("Emission Files/emission_report - 2017 to 2035.xlsx")

wb_2020_2030 = openpyxl.load_workbook("Emission Files/emission_report - 2020 to 2030.xlsx")
wb_2020_2035 = openpyxl.load_workbook("Emission Files/emission_report - 2020 to 2035.xlsx")

gdf = get_province_gdf()

pollutant_max = {
    'SO2': 1_000_000,
    'NOx': 2_500_000,
    'VOC': 3_000_000,
    'NH3': 1_000_000,
    'PM25': 750_000
}

pollutant_ticks = {
    "SO2": [0, 250000, 500000, 1000000],
    "NOx": [0, 1000000, 2000000, 2500000],
    "VOC": [0, 1000000, 2000000, 3000000],
    "NH3": [0, 400000, 800000, 1000000],
    "PM25": [0, 300000, 500000, 750000]
}

pollutant_ticklabels = {
    "SO2": ["0", "250k", "500k", "1M"],
    "NOx": ["0", "1M", "2M", "2.5M"],
    "VOC": ["0", "1M", "2M", "3M"],
    "NH3": ["0", "400k", "800k", "1M"],
    "PM25": ["0", "300k", "500k", "750k"]
}

for pollutant in ["SO2", "NOx", "VOC", "NH3", "PM25"]:
    cmap = plt.cm.YlOrRd
    norm = Normalize(vmin = 0, vmax = pollutant_max[pollutant])    
    
    def get_color_2017_2030(name):
        sheet = wb_2017_2030[pollutant]
        if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
            english_name = PROVINCE_MAP_EMISSIONS[name]
            province_index = PROVINCE_ARR.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"
    
    def get_color_2017_2035(name):
        sheet = wb_2017_2035[pollutant]
        if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
            english_name = PROVINCE_MAP_EMISSIONS[name]
            province_index = PROVINCE_ARR.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"

    def get_color_2020_2030(name):
        sheet = wb_2020_2030[pollutant]
        if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
            english_name = PROVINCE_MAP_EMISSIONS[name]
            province_index = PROVINCE_ARR.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"
    
    def get_color_2020_2035(name):
        sheet = wb_2020_2035[pollutant]
        if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
            english_name = PROVINCE_MAP_EMISSIONS[name]
            province_index = PROVINCE_ARR.index(english_name)
            total = 0
            for r in range(9, 14):
                total += sheet.cell(r, province_index + 5).value
            return cmap(norm(total))
        return "lightgrey"
    
    def get_color_2030(name):
        sheet = wb_baseline_2030[pollutant]
        if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
            english_name = PROVINCE_MAP_EMISSIONS[name]
            province_index = PROVINCE_ARR.index(english_name)
            total = 0
            for r in range(3, 8):
                total += sheet.cell(r, province_index + 3).value
            return cmap(norm(total))
        return "lightgrey"

    def get_color_2035(name):
        sheet = wb_baseline_2035[pollutant]
        if name in PROVINCE_MAP_EMISSIONS and name not in EXCLUDED_REGIONS:
            english_name = PROVINCE_MAP_EMISSIONS[name]
            province_index = PROVINCE_ARR.index(english_name)
            total = 0
            for r in range(3, 8):
                total += sheet.cell(r, province_index + 3).value
            return cmap(norm(total))
        return "lightgrey"
    
    fig, ax = plt.subplots(nrows = 3, ncols = 2, figsize = (30, 25), subplot_kw = {'projection': ccrs.PlateCarree()})
    
    gdf['color'] = gdf['province'].apply(get_color_2030)
    gdf.plot(color = gdf['color'], ax = ax[0][0], edgecolor = 'black', linewidth = 0.1)
    ax[0][0].set_title("2030 Emissions from DPEC")
    total = get_national_total(wb_baseline_2030, pollutant, range(3, 8), 3)
    ax[0][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[0][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2035)
    gdf.plot(color = gdf['color'], ax = ax[0][1], edgecolor = 'black', linewidth = 0.1)
    ax[0][1].set_title("2035 Emissions from DPEC")
    total = get_national_total(wb_baseline_2035, pollutant, range(3, 8), 3)
    ax[0][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[0][1].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2017_2030)
    gdf.plot(color = gdf['color'], ax = ax[1][0], edgecolor = 'black', linewidth = 0.1)
    ax[1][0].set_title("2030 Emissions Relative to 2017")
    total = get_national_total(wb_2017_2030, pollutant, range(9, 14), 5)
    ax[1][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[1][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2017_2035)
    gdf.plot(color = gdf['color'], ax = ax[1][1], edgecolor = 'black', linewidth = 0.1)
    ax[1][1].set_title("2035 Emissions Relative to 2017")
    total = get_national_total(wb_2017_2035, pollutant, range(9, 14), 5)
    ax[1][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[1][1].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2020_2030)
    gdf.plot(color = gdf['color'], ax = ax[2][0], edgecolor = 'black', linewidth = 0.1)
    ax[2][0].set_title("2030 Emissions Relative to 2020")
    total = get_national_total(wb_2020_2030, pollutant, range(9, 14), 5)
    ax[2][0].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[2][0].transAxes, ha = 'center', fontsize = 10)
    
    gdf['color'] = gdf['province'].apply(get_color_2020_2035)
    gdf.plot(color = gdf['color'], ax = ax[2][1], edgecolor = 'black', linewidth = 0.1)
    ax[2][1].set_title("2035 Emissions Relative to 2020")
    total = get_national_total(wb_2020_2035, pollutant, range(9, 14), 5)
    ax[2][1].text(0.5, 0.05, f"National Total: {total:,.0f} tons", 
                  transform = ax[2][1].transAxes, ha = 'center', fontsize = 10)
            
    plt.subplots_adjust(hspace = 0.1, wspace = 0.05)
    
    sm = cm.ScalarMappable(cmap = cmap, norm = norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax = ax, orientation = 'vertical', pad = 0.02)
    cbar.set_label('Absolute Emission Amount (tons)', fontsize = 14)
    cbar.set_ticks(pollutant_ticks[pollutant])
    cbar.set_ticklabels(pollutant_ticklabels[pollutant])

    fig.suptitle(f'Emission Amounts for {pollutant}', fontsize = 18, fontweight = 'bold')
    plt.savefig(f"Emission Files/Absolute Emissions Per Province/{pollutant}_Map.png")
    plt.close(fig)
    
print("Finished!")