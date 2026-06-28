import rasterio
import rasterio.windows
import netCDF4 as nc
import numpy as np
import xarray as xr
import geopandas as gpd
from shapely.vectorized import contains
from pathlib import Path
import openpyxl
from rasterio.features import rasterize
 
TIF_PATH = "Population Projection Gridded 2/SSP2RCP4_5/grid_pop_count2035_SSP2_RCP4_5.tif"           
OUT_PATH = "Emission Files/pop_grid_wrf_2035.npy"                       
HALF_WRF_DEG = 0.18
 
ds = nc.Dataset("2017_Base_Data/PM25.nc")
lat = ds.variables["lat"][:]   
lon = ds.variables["lon"][:]  
ny, nx = lat.shape
 
pop_grid = np.full((ny, nx), np.nan, dtype = np.float64)
 
with rasterio.open(TIF_PATH) as src:
    nodata = src.nodata
 
    for i in range(ny):
        for j in range(nx):
            clat = float(lat[i, j])
            clon = float(lon[i, j])
 
            left = clon - HALF_WRF_DEG
            right = clon + HALF_WRF_DEG
            bottom = clat - HALF_WRF_DEG
            top = clat + HALF_WRF_DEG
 
            row_start, col_start = src.index(left, top)
            row_stop,  col_stop  = src.index(right, bottom)
 
            row_start = max(0, row_start)
            col_start = max(0, col_start)
            row_stop  = min(src.height, row_stop)
            col_stop  = min(src.width,  col_stop)
 
            if row_stop <= row_start or col_stop <= col_start:
                continue
 
            window = rasterio.windows.Window(
                col_start, row_start,
                col_stop - col_start,
                row_stop - row_start,
            )
            
            data = src.read(1, window = window).astype(np.float64)
            data[data == nodata] = np.nan
 
            valid = data[~np.isnan(data)]
            if len(valid) > 0:
                pop_grid[i, j] = valid.sum()
 
        if i % 20 == 0:
            print(f"  Row {i}/{ny} complete…")
            
np.save(OUT_PATH, pop_grid)
            
world = gpd.read_file("https://naturalearth.s3.amazonaws.com/110m_cultural/ne_110m_admin_0_countries.zip")
china = world[world['NAME'] == 'China']

lon_flat = ds["lon"][:].flatten()
lat_flat = ds["lat"][:].flatten()

in_china_flat = contains(china.geometry.iloc[0], lon_flat, lat_flat)

in_china = in_china_flat.reshape(ds["lon"].shape)

pop_grid_2035 = np.load("Emission Files/pop_grid_wrf_2035.npy")
pop_china_2035 = pop_grid_2035.copy()
pop_china_2035 = np.nan_to_num(pop_china_2035, nan = 0.0)
            
wb = openpyxl.Workbook()
sheet = wb.active
            
nc_files = sorted(Path("NC Files and Emission Reports").glob("*.nc"))
for file in nc_files:
    ds_2035 = xr.open_dataset(file)
    data = ds_2035["pred_PM25"].mean(dim = "time").values
    
    above_25 = (data >= 25)
    mask = in_china & above_25
    
    print(f"{file.stem}: {data[in_china].mean()}")
    
    percentage = round((pop_china_2035[mask].sum() / pop_china_2035[in_china].sum()) * 100, 2)
    
    parts = file.stem.split("_")
    scenario = parts[1]
    year = parts[2].split("Met")[0]
    
    sheet.append([scenario, year, percentage])
    
wb.save("Emission Files/2035 Scenarios above 25 Pop. Percentage.xlsx")
print("Finished Part 1!")

TIF_PATH = "Population Projection Gridded 2/SSP2RCP4_5/grid_pop_count2035_SSP2_RCP4_5.tif"

with rasterio.open(TIF_PATH) as src:
    pop_1km = src.read(1).astype('float64')
    transform = src.transform
    nodata = src.nodata
    shape = pop_1km.shape
    crs = src.crs
    
pop_1km[pop_1km == nodata] = 0

print(f"Shape: {shape}")
print(f"Nodata: {nodata}")
print(f"Total pop: {pop_1km[pop_1km != nodata].sum():,.0f}")

province_map = gpd.read_file("https://geo.datav.aliyun.com/areas_v3/bound/100000_full.json")
province_map = province_map[['name', 'geometry']].rename(columns = {'name': 'province'})
province_map = province_map.to_crs(crs)

province_map['id'] = range(1, len(province_map) + 1)
id_to_province = dict(zip(province_map['id'], province_map['province']))

province_raster = rasterize(
    [(geom, pid) for geom, pid in zip(province_map.geometry, province_map['id'])],
    out_shape = shape,
    transform = transform,
    fill = 0,
    dtype = 'int32'
)

ds_ref = nc.Dataset("2017_Base_Data/PM25.nc")
wrf_lat = ds_ref.variables["lat"][:]
wrf_lon = ds_ref.variables["lon"][:]
ny, nx = wrf_lat.shape

HALF_WRF_DEG = 0.18

with rasterio.open(TIF_PATH) as src:
    wrf_raster = np.full(shape, -1, dtype = np.int32)
    
    for i in range(ny):
        for j in range(nx):
            clat = float(wrf_lat[i, j])
            clon = float(wrf_lon[i, j])

            left = clon - HALF_WRF_DEG
            right = clon + HALF_WRF_DEG
            bottom = clat - HALF_WRF_DEG
            top = clat + HALF_WRF_DEG

            row_start, col_start = src.index(left, top)
            row_stop,  col_stop = src.index(right, bottom)

            row_start = max(0, row_start)
            col_start = max(0, col_start)
            row_stop  = min(src.height, row_stop)
            col_stop  = min(src.width,  col_stop)

            if row_stop <= row_start or col_stop <= col_start:
                continue

            wrf_raster[row_start : row_stop, col_start : col_stop] = i * nx + j

        if i % 20 == 0:
            print(f"WRF raster row {i}/{ny}")

np.save("Emission Files/province_raster_2035.npy", province_raster)
np.save("Emission Files/wrf_raster_2035.npy", wrf_raster)

province_raster = np.load("Emission Files/province_raster_2035.npy")
wrf_raster = np.load("Emission Files/wrf_raster_2035.npy")

wb = openpyxl.Workbook()
ws = wb.active
for col_index, pname in enumerate(id_to_province.values(), 1):
        ws.cell(1, col_index).value = pname

nc_files = sorted(Path("NC Files and Emission Reports").glob("*.nc"))
for row_index, file in enumerate(nc_files, 2):
    pm25 = xr.open_dataset(file)["pred_PM25"].mean(dim = "time").values
    pm25_flat = pm25.flatten()
    pm25_1km = pm25_flat[wrf_raster]
    pm25_1km[wrf_raster == -1] = np.nan
    
    above_25 = pm25_1km >= 25

    col_index = 1
    for pid, pname in id_to_province.items():
        prov_mask = province_raster == pid
        total = pop_1km[prov_mask].sum()
        if total > 0:
            exposed = pop_1km[prov_mask & above_25].sum()
            pct = round((exposed / total) * 100, 2)
            ws.cell(row_index, col_index).value = round(pct, 2)
        col_index = col_index + 1

wb.save("Emission Files/Province Population Percentage 2035.xlsx")
print("Done!")