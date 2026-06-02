import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\early_peak-net_zero-clean_air\2020_2035_change_percentage.xlsx")

wb_edit = openpyxl.load_workbook(r"C:\Users\sunyi\Environment_Project\Emission Excel Files\管控方案模板.xlsx")
ws_edit = wb_edit["管控方案"]
ws_edit.delete_rows(3, 10000)

for sheet_name in ["SO2", "NOX", "VOC", "NH3", "PM2.5"]:
    ws = wb[sheet_name]
    for r in range(2, 7):
        for c in range(2, 33):
            province = ws.cell(row = 1, column = c).value
            department = ws.cell(row = r, column = 1).value
            value = ws.cell(row = r, column = c).value

            if value == "#DIV/0!":
                value = 0
            elif float(value) < 0:
                value = 0
                
            if sheet_name == "PM2.5":
                sheet_name = "PM"

            ws_edit.append([province, department, sheet_name, value])


for weather_year in range(2017, 2025):
    wb_edit.save(rf"C:\Users\sunyi\Environment_Project\Scenario Excel Files\early_peak-net_zero-clean_air-2020-{weather_year}_2020_{weather_year}.xlsx")

print("Finished!")
