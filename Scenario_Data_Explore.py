import xarray as xr
import geopandas as gpd
import shapely
from pathlib import Path
import zipfile
import openpyxl

wb = openpyxl.Workbook()
wb.remove(wb.active)

file_path = Path(r"C:\Users\sunyi\Environment_Project\Excel Downloads")

list_zips = file_path.glob("*.zip")

world = gpd.read_file("https://naturalearth.s3.amazonaws.com/110m_cultural/ne_110m_admin_0_countries.zip")
china = world[world['NAME'] == 'China']

abbrev = {
    "baseline": "base",
    "clean-air": "ca",
    "early-peak-net-zero-clean-air": "epnzca",
    "on-time-peak-clean-air": "otpca",
    "on-time-peak-net-zero-clean-air": "otpnzca"
}

count = 1
for file_path in list_zips:
    with zipfile.ZipFile(file_path, "r") as z:
        with z.open("PM25.nc") as src:
            base_name = Path(file_path).stem.split("-2020-")
            zip_name = abbrev.get(base_name[0])
            wb.create_sheet(title = f"{zip_name}-{base_name[1]}")
            output_path = Path(file_path).parent.parent / "NC Files" / f"{zip_name}-{base_name[1]}.nc"
            with open(output_path, "wb") as dst:
                dst.write(src.read())
                
    ds = xr.open_dataset(output_path)
    
    lon_flat = ds["lon"].values.flatten()
    lat_flat = ds["lat"].values.flatten()
    
    in_china_flat = shapely.contains_xy(china.geometry.iloc[0], lon_flat, lat_flat)
    in_china = in_china_flat.reshape(ds["lon"].shape)
    
    day_data = ds["pred_PM25"].mean(dim="time").values
    china_cells = day_data[in_china]
    pct = (china_cells >= 25).sum() / len(china_cells) * 100
    
    wb[f"{zip_name}-{base_name[1]}"].append(["Annual Mean", pct])
    
    for day in range(365):
        day_data = ds["pred_PM25"][day].values
        china_cells = day_data[in_china]
        pct = (china_cells >= 25).sum() / len(china_cells) * 100
        wb[f"{zip_name}-{base_name[1]}"].append([f"Day {day + 1}", pct])
    
        print(f"{count} {zip_name}-{base_name[1]} Day {day + 1}: {pct:.2f}% of China cells >= 25")
    
    count += 1
    
print(wb.sheetnames)
    
wb.save(r"C:\Users\sunyi\Environment_Project\Emission Excel Files\Percent_China_Over_25.xlsx")
print("Done!")